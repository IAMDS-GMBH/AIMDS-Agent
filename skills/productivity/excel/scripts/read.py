#!/usr/bin/env python3
"""Read and analyse Excel (.xlsx / .xls / .csv) files.

Usage:
    python read.py workbook.xlsx                       # first 200 rows of the first sheet
    python read.py workbook.xlsx --sheet "Q2"          # specific sheet
    python read.py workbook.xlsx --max-rows 50         # cap rows printed (default 200)
    python read.py workbook.xlsx --range A1:D20        # exact cell range (TSV)
    python read.py workbook.xlsx --sheets              # list all sheet names
    python read.py workbook.xlsx --stats               # summary statistics
    python read.py workbook.xlsx --cell B5             # read a single cell
    python read.py workbook.xlsx --metadata
"""

import json
import sys
from pathlib import Path

DEFAULT_MAX_ROWS = 200


def _is_csv(path: str | Path) -> bool:
    return Path(path).suffix.lower() == ".csv"


def list_sheets(path):
    if _is_csv(path):
        print("  Sheet1  (CSV source)")
        return

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name}  ({ws.max_row} rows × {ws.max_column} cols)")
    wb.close()


def _load_frame(path, sheet=None, nrows=None):
    import pandas as pd

    if _is_csv(path):
        return pd.read_csv(path, nrows=nrows)
    return pd.read_excel(path, sheet_name=sheet or 0, nrows=nrows)


def read_sheet(path, sheet=None, max_rows=DEFAULT_MAX_ROWS):
    max_rows = max(1, int(max_rows))
    # Read one row past the cap so we can tell the caller the output is truncated
    # without loading a huge sheet into memory.
    df = _load_frame(path, sheet, nrows=max_rows + 1)
    truncated = len(df) > max_rows
    if truncated:
        df = df.head(max_rows)
    print(df.to_string(index=False))
    if truncated:
        print(
            f"\n[truncated] Showing the first {max_rows} data rows only. "
            "Raise --max-rows or use --range A1:D50 to read a specific block."
        )


def read_range(path, cell_range, sheet=None):
    if _is_csv(path):
        print("Range lookup is not supported for CSV input. Convert to .xlsx first.", file=sys.stderr)
        sys.exit(1)

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        cells = ws[cell_range]
        if not isinstance(cells, tuple):
            cells = ((cells,),)
        elif cells and not isinstance(cells[0], tuple):
            cells = (cells,)
        rows = 0
        for row in cells:
            print("\t".join("" if c.value is None else str(c.value) for c in row))
            rows += 1
        print(f"\n[{rows} row(s) from {ws.title}!{cell_range}]")
    finally:
        wb.close()


def show_stats(path, sheet=None):
    df = _load_frame(path, sheet)
    print(df.describe(include="all").to_string())


def read_cell(path, cell_ref, sheet=None):
    if _is_csv(path):
        print("Single-cell lookup is not supported for CSV input. Convert to .xlsx first.")
        sys.exit(1)

    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    print(ws[cell_ref].value)
    wb.close()


def show_metadata(path):
    if _is_csv(path):
        import pandas as pd

        df = pd.read_csv(path)
        print(
            json.dumps(
                {
                    "title": "",
                    "subject": "",
                    "creator": "",
                    "keywords": "",
                    "created": "",
                    "modified": "",
                    "sheets": ["Sheet1"],
                    "rows": int(len(df)),
                    "columns": int(len(df.columns)),
                },
                indent=2,
            )
        )
        return

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    props = wb.properties
    print(
        json.dumps(
            {
                "title": props.title or "",
                "subject": props.subject or "",
                "creator": props.creator or "",
                "keywords": props.keywords or "",
                "created": str(props.created or ""),
                "modified": str(props.modified or ""),
                "sheets": wb.sheetnames,
                "dimensions": {name: f"{wb[name].max_row}x{wb[name].max_column}" for name in wb.sheetnames},
            },
            indent=2,
        )
    )
    wb.close()


def _opt(args, flag, default=None):
    if flag in args:
        idx = args.index(flag)
        if idx + 1 < len(args):
            return args[idx + 1]
    return default


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] in {"-h", "--help"}:
        print(__doc__)
        sys.exit(0)

    path = args[0]
    sheet = _opt(args, "--sheet")

    if "--sheets" in args:
        list_sheets(path)
    elif "--stats" in args:
        show_stats(path, sheet)
    elif "--cell" in args:
        read_cell(path, _opt(args, "--cell"), sheet)
    elif "--range" in args:
        read_range(path, _opt(args, "--range"), sheet)
    elif "--metadata" in args:
        show_metadata(path)
    else:
        read_sheet(path, sheet, max_rows=_opt(args, "--max-rows", DEFAULT_MAX_ROWS))
