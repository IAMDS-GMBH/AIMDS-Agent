#!/usr/bin/env python3
"""Create and edit Excel (.xlsx) workbooks.

Usage:
    # Create an empty workbook (optionally with named sheets)
    python write.py --new report.xlsx --sheets "Data,Summary"

    # Create a workbook from a CSV file
    python write.py --from-csv data.csv output.xlsx

    # Update one cell
    python write.py --set-cell B5 "=SUM(B2:B4)" workbook.xlsx
    python write.py --set-cell B5 "=SUM(B2:B4)" workbook.xlsx --sheet "Q2"

    # Update many cells in one pass (JSON object cell->value, or list of [cell, value])
    python write.py --set-cells '{"A1": "Region", "B1": "Revenue", "B2": 1200.5}' workbook.xlsx

    # Append a row (real CSV parsing: quotes protect commas) or JSON values
    python write.py --append-row 'April,"1,200",85000' workbook.xlsx
    python write.py --append-json '["April", 1200, 85000]' workbook.xlsx

    # Format a range (bold, fill, number_format, align, wrap, col_width, row_height, autofit)
    python write.py --format '{"range": "A1:C1", "bold": true, "fill": "3F59FF", "font_color": "FFFFFF"}' workbook.xlsx
    python write.py --format '{"range": "B2:B50", "number_format": "#,##0.00", "autofit": true}' workbook.xlsx

    # Export a sheet to CSV
    python write.py --to-csv workbook.xlsx output.csv
    python write.py --to-csv workbook.xlsx output.csv --sheet "Q2"

    # Convert to PDF via LibreOffice
    python write.py --to-pdf workbook.xlsx
"""

import csv
import io
import json
import shutil
import subprocess
import sys
from pathlib import Path


def _libreoffice_cmd() -> str | None:
    for candidate in [
        "libreoffice",
        "soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]:
        if "/" in candidate:
            if Path(candidate).exists():
                return candidate
            continue
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    return None


def _coerce(value):
    """Best-effort scalar coercion for CLI strings: int, float, bool, else str.

    Formulas (leading ``=``) and anything that does not parse stay strings.
    """
    if not isinstance(value, str):
        return value
    v = value.strip()
    if v.startswith("="):
        return v
    if v.lower() in {"true", "false"}:
        return v.lower() == "true"
    try:
        return int(v)
    except ValueError:
        pass
    try:
        return float(v)
    except ValueError:
        return value


def _open(path, sheet=None):
    import openpyxl

    wb = openpyxl.load_workbook(path)
    if sheet:
        if sheet not in wb.sheetnames:
            print(f"Error: sheet {sheet!r} not found. Available: {wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
        ws = wb[sheet]
    else:
        ws = wb.active
    return wb, ws


def new_workbook(dst, sheets=None):
    import openpyxl

    wb = openpyxl.Workbook()
    names = [s.strip() for s in (sheets or "").split(",") if s.strip()]
    if names:
        wb.active.title = names[0]
        for name in names[1:]:
            wb.create_sheet(name)
    Path(dst).parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    print(f"Created {dst} with sheets {wb.sheetnames}")


def from_csv(src, dst):
    import pandas as pd

    df = pd.read_csv(src)
    df.to_excel(dst, index=False)
    print(f"Created {dst} from {src} ({len(df)} rows)")


def _parse_cells(cells_json):
    data = json.loads(cells_json)
    if isinstance(data, dict):
        return list(data.items())
    if isinstance(data, list):
        items = []
        for entry in data:
            if isinstance(entry, dict) and "cell" in entry:
                items.append((entry["cell"], entry.get("value")))
            elif isinstance(entry, (list, tuple)) and len(entry) == 2:
                items.append((entry[0], entry[1]))
            else:
                raise ValueError(f"Unsupported cell entry: {entry!r}")
        return items
    raise ValueError("cells JSON must be an object {cell: value} or a list of [cell, value]")


def set_cells(path, cells_json, sheet=None):
    items = _parse_cells(cells_json)
    wb, ws = _open(path, sheet)
    for cell_ref, value in items:
        ws[cell_ref] = _coerce(value) if isinstance(value, str) and value.startswith("=") else value
    wb.save(path)
    print(f"Set {len(items)} cell(s) in {ws.title} of {path}")
    print(json.dumps({"cells_set": len(items), "sheet": ws.title, "file": str(path)}))


def set_cell(path, cell_ref, value, sheet=None):
    set_cells(path, json.dumps({cell_ref: _coerce(value)}), sheet)


def append_row(path, row_csv=None, values=None, sheet=None):
    if values is None:
        parsed = next(csv.reader(io.StringIO(row_csv or "")))
        values = [_coerce(v) for v in parsed]
    wb, ws = _open(path, sheet)
    ws.append(values)
    wb.save(path)
    print(f"Appended row {ws.max_row} to {ws.title} of {path}: {values}")


def _autofit(ws, cells):
    from openpyxl.utils import get_column_letter

    widths: dict[int, int] = {}
    for row in cells:
        for c in row:
            if c.value is None:
                continue
            length = max(len(line) for line in str(c.value).splitlines() or [""])
            widths[c.column] = max(widths.get(c.column, 0), length)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 8), 80)


def format_cells(path, spec_json, sheet=None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    spec = json.loads(spec_json)
    if not isinstance(spec, dict) or not spec.get("range"):
        print("Error: format spec must be a JSON object with a 'range' key", file=sys.stderr)
        sys.exit(1)

    wb, ws = _open(path, sheet)
    cells = ws[spec["range"]]
    if not isinstance(cells, tuple):
        cells = ((cells,),)
    elif cells and not isinstance(cells[0], tuple):
        cells = (cells,)

    font_kwargs = {}
    for key in ("bold", "italic"):
        if key in spec:
            font_kwargs[key] = bool(spec[key])
    if spec.get("font_color"):
        font_kwargs["color"] = str(spec["font_color"]).lstrip("#").upper()
    if spec.get("font_size"):
        font_kwargs["size"] = float(spec["font_size"])
    if spec.get("font_name"):
        font_kwargs["name"] = str(spec["font_name"])

    fill = None
    if spec.get("fill"):
        color = str(spec["fill"]).lstrip("#").upper()
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    align_kwargs = {}
    if spec.get("align"):
        align_kwargs["horizontal"] = str(spec["align"])
    if spec.get("valign"):
        align_kwargs["vertical"] = str(spec["valign"])
    if "wrap" in spec:
        align_kwargs["wrap_text"] = bool(spec["wrap"])

    touched = 0
    for row in cells:
        for c in row:
            if font_kwargs:
                base = c.font
                c.font = Font(
                    name=font_kwargs.get("name", base.name),
                    size=font_kwargs.get("size", base.size),
                    bold=font_kwargs.get("bold", base.bold),
                    italic=font_kwargs.get("italic", base.italic),
                    color=font_kwargs.get("color", base.color),
                )
            if fill is not None:
                c.fill = fill
            if spec.get("number_format"):
                c.number_format = str(spec["number_format"])
            if align_kwargs:
                c.alignment = Alignment(**align_kwargs)
            touched += 1

    col_width = spec.get("col_width")
    if isinstance(col_width, dict):
        for letter, width in col_width.items():
            ws.column_dimensions[str(letter).upper()].width = float(width)
    elif col_width is not None:
        for col in range(cells[0][0].column, cells[0][-1].column + 1):
            ws.column_dimensions[get_column_letter(col)].width = float(col_width)
    if spec.get("row_height") is not None:
        for row in cells:
            ws.row_dimensions[row[0].row].height = float(spec["row_height"])
    if spec.get("autofit"):
        _autofit(ws, cells)

    wb.save(path)
    print(f"Formatted {touched} cell(s) in {ws.title}!{spec['range']} of {path}")


def to_csv(path, out, sheet=None):
    import pandas as pd

    df = pd.read_excel(path, sheet_name=sheet or 0)
    df.to_csv(out, index=False, encoding="utf-8-sig")
    print(f"Exported {out} ({len(df)} rows)")


def to_pdf(path):
    path = str(Path(path).resolve())
    out = Path(path).with_suffix(".pdf")
    libreoffice_bin = _libreoffice_cmd()
    if not libreoffice_bin:
        print(
            "Error: LibreOffice/soffice not found. Install LibreOffice to enable XLSX -> PDF conversion.",
            file=sys.stderr,
        )
        sys.exit(1)

    result = subprocess.run(
        [
            libreoffice_bin,
            "--headless",
            "--convert-to",
            "pdf",
            str(path),
            "--outdir",
            str(Path(path).parent),
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print(f"Error: {result.stderr}", file=sys.stderr)
        sys.exit(1)
    if not out.exists():
        print("Error: Conversion did not produce a PDF output file.", file=sys.stderr)
        sys.exit(1)
    print(f"PDF: {out}")


def _opt(args, flag, default=None):
    if flag in args:
        idx = args.index(flag)
        if idx + 1 < len(args):
            return args[idx + 1]
    return default


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] in {"-h", "--help"}:
        print(__doc__)
        sys.exit(0)

    sheet = _opt(args, "--sheet")

    if "--new" in args:
        new_workbook(_opt(args, "--new"), _opt(args, "--sheets"))

    elif "--from-csv" in args:
        idx = args.index("--from-csv")
        from_csv(args[idx + 1], args[idx + 2])

    elif "--set-cell" in args:
        idx = args.index("--set-cell")
        set_cell(args[idx + 3], args[idx + 1], args[idx + 2], sheet)

    elif "--set-cells" in args:
        idx = args.index("--set-cells")
        set_cells(args[idx + 2], args[idx + 1], sheet)

    elif "--append-row" in args:
        idx = args.index("--append-row")
        append_row(args[idx + 2], row_csv=args[idx + 1], sheet=sheet)

    elif "--append-json" in args:
        idx = args.index("--append-json")
        values = json.loads(args[idx + 1])
        if not isinstance(values, list):
            print("Error: --append-json expects a JSON array", file=sys.stderr)
            sys.exit(1)
        append_row(args[idx + 2], values=values, sheet=sheet)

    elif "--format" in args:
        idx = args.index("--format")
        format_cells(args[idx + 2], args[idx + 1], sheet)

    elif "--to-csv" in args:
        idx = args.index("--to-csv")
        path = str(Path(args[idx + 1]).resolve())
        out = (
            args[idx + 2]
            if idx + 2 < len(args) and not args[idx + 2].startswith("--")
            else Path(path).with_suffix(".csv")
        )
        to_csv(path, str(out), sheet)

    elif "--to-pdf" in args:
        to_pdf(_opt(args, "--to-pdf"))

    else:
        print("Unknown arguments. Run with --help for usage.")
        sys.exit(1)
